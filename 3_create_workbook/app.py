from openpyxl import Workbook
wb = Workbook()

# grab the active worksheet
ws = wb.active

#or
#ws2 = wb["Sheet2"]

ws1 = wb.create_sheet("Mysheet") # insert at the end (default)
#or
ws2 = wb.create_sheet("Mysheet", 0) # insert at first position
#or
ws3 = wb.create_sheet("Mysheet", -1) # insert at the penultimate position

wb.save("sample3.xlsx")